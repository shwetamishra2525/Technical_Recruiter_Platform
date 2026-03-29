import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta

# Create workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt Chart (2025-2026)"

# Define initial headers
headers = ["Task Name", "Duration (Days)", "Start Date", "Finish Date"]
for col, val in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=val)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15

# Define the project timeline (Oct 1, 2025 to Mar 31, 2026) -> 182 days
project_start = datetime(2025, 10, 1)
project_end = datetime(2026, 3, 31)
total_days = (project_end - project_start).days + 1

# Generate day columns in Excel (Column E onwards)
for i in range(total_days):
    current_date = project_start + timedelta(days=i)
    col = 5 + i
    cell = ws.cell(row=1, column=col, value=current_date.strftime("%d-%b"))
    cell.font = Font(size=8, bold=True)
    cell.alignment = Alignment(text_rotation=90)
    # Make columns very narrow for Gantt chart feel
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 3

# Data structure: (Task Name, Start, End, IsHeader)
tasks = [
    ("Technical Recruiter Platform", "2025-10-01", "2026-03-31", True),
    ("Requirement Gathering", "2025-10-01", "2025-10-10", False),
    ("Planning", "2025-10-11", "2025-10-20", False),
    ("Analysis and Design", "2025-10-21", "2025-10-31", False),
    
    ("Module 1: Registration & Auth", "2025-11-01", "2025-11-25", True),
    ("  Planning", "2025-11-01", "2025-11-05", False),
    ("  Analysis and Design", "2025-11-06", "2025-11-10", False),
    ("  Coding", "2025-11-11", "2025-11-20", False),
    ("  Testing", "2025-11-21", "2025-11-25", False),
    
    ("Module 2: Job Descriptions", "2025-11-26", "2025-12-20", True),
    ("  Planning", "2025-11-26", "2025-11-30", False),
    ("  Analysis and Design", "2025-12-01", "2025-12-05", False),
    ("  Coding", "2025-12-06", "2025-12-15", False),
    ("  Testing", "2025-12-16", "2025-12-20", False),
    
    ("Module 3: AI Resume Parsing", "2025-12-21", "2026-02-05", True),
    ("  Planning", "2025-12-21", "2025-12-27", False),
    ("  Analysis and Design", "2025-12-28", "2026-01-03", False),
    ("  Coding", "2026-01-04", "2026-01-28", False),
    ("  Testing", "2026-01-29", "2026-02-05", False),

    ("Module 4: Interview Room", "2026-02-06", "2026-03-18", True),
    ("  Planning", "2026-02-06", "2026-02-12", False),
    ("  Analysis and Design", "2026-02-13", "2026-02-19", False),
    ("  Coding", "2026-02-20", "2026-03-11", False),
    ("  Testing", "2026-03-12", "2026-03-18", False),
    
    ("Final Integration & Deployment", "2026-03-19", "2026-03-31", True),
    ("  System Testing", "2026-03-19", "2026-03-24", False),
    ("  Bug Fixing", "2026-03-25", "2026-03-28", False),
    ("  Deployment & Report", "2026-03-29", "2026-03-31", False)
]

# Write tasks and draw Gantt bars
fill_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
fill_bar = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid") # MS Project Blueish

row_idx = 2
for task in tasks:
    name, start_str, end_str, is_header = task
    
    start_dt = datetime.strptime(start_str, "%Y-%m-%d")
    end_dt = datetime.strptime(end_str, "%Y-%m-%d")
    duration = (end_dt - start_dt).days + 1
    
    # Write details
    ws.cell(row=row_idx, column=1, value=name)
    ws.cell(row=row_idx, column=2, value=duration)
    ws.cell(row=row_idx, column=3, value=start_dt.strftime("%d-%b-%y"))
    ws.cell(row=row_idx, column=4, value=end_dt.strftime("%d-%b-%y"))
    
    if is_header:
        for i in range(1, 5):
            ws.cell(row=row_idx, column=i).font = Font(bold=True)
            ws.cell(row=row_idx, column=i).fill = fill_header
            
    # Draw bars by coloring cells!
    start_offset = (start_dt - project_start).days
    for i in range(duration):
        col = 5 + start_offset + i
        if not is_header:
            ws.cell(row=row_idx, column=col).fill = fill_bar
        else:
            # Maybe a thin grey bar for headers to mimic MS Project Summary Tasks
            ws.cell(row=row_idx, column=col).fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    row_idx += 1

# Save file
wb.save(r"e:\MSC-II\Technical_Recruiter Project\Technical_Recruiter_Gantt_Chart.xlsx")
print("Saved Technical_Recruiter_Gantt_Chart.xlsx successfully!")
