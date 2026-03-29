import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt Chart (Aug - Mar)"

# Set column widths for textual data
ws.column_dimensions['A'].width = 35  
ws.column_dimensions['B'].width = 12  
ws.column_dimensions['C'].width = 18  
ws.column_dimensions['D'].width = 18  

headers = ["Task Name", "Duration", "Start", "Finish"]
for col, val in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col, value=val)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Project timeline (Aug 1 to Mar 20)
project_start = datetime(2025, 8, 1)
project_end = datetime(2026, 3, 31) # Render up to Mar 31 for grid
total_days = (project_end - project_start).days + 1

for i in range(total_days):
    col = 5 + i
    col_letter = openpyxl.utils.get_column_letter(col)
    ws.column_dimensions[col_letter].width = 0.7  # slightly thinner since it's 8 months

months = [
    ("Aug 2025", 31), ("Sep 2025", 30), ("Oct 2025", 31), ("Nov 2025", 30),
    ("Dec 2025", 31), ("Jan 2026", 31), ("Feb 2026", 28), ("Mar 2026", 31)
]

current_col = 5
for m_name, d_count in months:
    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + d_count - 1)
    cell = ws.cell(row=1, column=current_col, value=m_name)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    
    thin_border = Border(right=Side(style='thin', color="C0C0C0"))
    for r in range(1, 35):
        ws.cell(row=r, column=current_col + d_count - 1).border = thin_border
        
    current_col += d_count

# Tasks Data (Stretched from Aug 1 to Mar 20)
tasks = [
    ("Technical Recruiter Platform", "2025-08-01", "2026-03-20", True, "232 days"),
    ("Requirement Gathering", "2025-08-01", "2025-08-20", False, "20 days"),
    ("Planning", "2025-08-21", "2025-09-10", False, "21 days"),
    ("Analysis and Design", "2025-09-11", "2025-09-30", False, "20 days"),
    
    ("Module 1: Registration & Auth", "2025-10-01", "2025-10-31", True, "1 mon"),
    ("    Planning", "2025-10-01", "2025-10-05", False, "5 days"),
    ("    Analysis and Design", "2025-10-06", "2025-10-10", False, "5 days"),
    ("    Coding", "2025-10-11", "2025-10-25", False, "15 days"),
    ("    Testing", "2025-10-26", "2025-10-31", False, "6 days"),
    
    ("Module 2: Job Descriptions", "2025-11-01", "2025-11-30", True, "1 mon"),
    ("    Planning", "2025-11-01", "2025-11-05", False, "5 days"),
    ("    Analysis and Design", "2025-11-06", "2025-11-10", False, "5 days"),
    ("    Coding", "2025-11-11", "2025-11-25", False, "15 days"),
    ("    Testing", "2025-11-26", "2025-11-30", False, "5 days"),
    
    ("Module 3: AI Resume Parsing", "2025-12-01", "2026-01-25", True, "1.8 mons"),
    ("    Planning", "2025-12-01", "2025-12-10", False, "10 days"),
    ("    Analysis and Design", "2025-12-11", "2025-12-20", False, "10 days"),
    ("    Coding", "2025-12-21", "2026-01-15", False, "26 days"),
    ("    Testing", "2026-01-16", "2026-01-25", False, "10 days"),

    ("Module 4: Interview Room", "2026-01-26", "2026-03-12", True, "1.5 mons"),
    ("    Planning", "2026-01-26", "2026-02-05", False, "11 days"),
    ("    Analysis and Design", "2026-02-06", "2026-02-15", False, "10 days"),
    ("    Coding", "2026-02-16", "2026-03-05", False, "18 days"),
    ("    Testing", "2026-03-06", "2026-03-12", False, "7 days"),
    
    ("Final Integration & Deployment", "2026-03-13", "2026-03-20", True, "8 days"),
    ("    System Testing", "2026-03-13", "2026-03-16", False, "4 days"),
    ("    Bug Fixing", "2026-03-17", "2026-03-18", False, "2 days"),
    ("    Deployment & Report", "2026-03-19", "2026-03-20", False, "2 days")
]

fill_bar = PatternFill(start_color="71C1C4", end_color="71C1C4", fill_type="solid") 
fill_header_bar = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")

row_idx = 3
for task in tasks:
    name, start_str, end_str, is_header, duration_str = task
    
    start_dt = datetime.strptime(start_str, "%Y-%m-%d")
    end_dt = datetime.strptime(end_str, "%Y-%m-%d")
    duration_days = (end_dt - start_dt).days + 1
    
    ws.cell(row=row_idx, column=1, value=name)
    ws.cell(row=row_idx, column=2, value=duration_str)
    ws.cell(row=row_idx, column=3, value=start_dt.strftime("%a %d-%m-%y"))
    ws.cell(row=row_idx, column=4, value=end_dt.strftime("%a %d-%m-%y"))
    
    if is_header:
        for i in range(1, 5):
            ws.cell(row=row_idx, column=i).font = Font(bold=True)
            
    start_offset = (start_dt - project_start).days
    for i in range(duration_days):
        col = 5 + start_offset + i
        if not is_header:
            ws.cell(row=row_idx, column=col).fill = fill_bar
        else:
            ws.cell(row=row_idx, column=col).fill = fill_header_bar

    if not is_header:
        border_style = Border(top=Side(style='thin', color="FFFFFF"), bottom=Side(style='thin', color="FFFFFF"))
        for i in range(duration_days):
            ws.cell(row=row_idx, column=5 + start_offset + i).border = border_style

    row_idx += 1

wb.save(r"e:\MSC-II\Technical_Recruiter Project\Technical_Recruiter_Gantt_Chart_v4.xlsx")
print("Saved v4 successfully!")
