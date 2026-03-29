import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt Chart (2025-2026)"

# Set column widths for textual data
ws.column_dimensions['A'].width = 35  # Task Name
ws.column_dimensions['B'].width = 12  # Duration
ws.column_dimensions['C'].width = 18  # Start
ws.column_dimensions['D'].width = 18  # Finish

# Headers for textual data (Row 2, because Row 1 will have Month Headers)
headers = ["Task Name", "Duration", "Start", "Finish"]
for col, val in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col, value=val)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Define the project timeline
project_start = datetime(2025, 10, 1)
project_end = datetime(2026, 3, 31)
total_days = (project_end - project_start).days + 1

# Make the day columns VERY thin so it all fits on one screen (Width = 0.8)
for i in range(total_days):
    col = 5 + i
    col_letter = openpyxl.utils.get_column_letter(col)
    ws.column_dimensions[col_letter].width = 0.8

# Month grouping for Row 1
months = [
    ("Oct 2025", 31), ("Nov 2025", 30), ("Dec 2025", 31),
    ("Jan 2026", 31), ("Feb 2026", 28), ("Mar 2026", 31)
]

current_col = 5
for m_name, d_count in months:
    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + d_count - 1)
    cell = ws.cell(row=1, column=current_col, value=m_name)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    
    # Optional: Put faint lines separating months
    thin_border = Border(right=Side(style='thin'))
    for r in range(1, 35):
        ws.cell(row=r, column=current_col + d_count - 1).border = thin_border
        
    current_col += d_count

# Tasks Data
# Format: (Name, StartDate, EndDate, IsHeader, InitialDurationOverrideString)
tasks = [
    ("Technical Recruiter Platform", "2025-10-01", "2026-03-31", True, "6 mons"),
    ("Requirement Gathering", "2025-10-01", "2025-10-10", False, "10 days"),
    ("Planning", "2025-10-11", "2025-10-20", False, "10 days"),
    ("Analysis and Design", "2025-10-21", "2025-10-31", False, "10 days"),
    
    ("Module 1: Registration & Auth", "2025-11-01", "2025-11-25", True, "1 mon"),
    ("    Planning", "2025-11-01", "2025-11-05", False, "5 days"),
    ("    Analysis and Design", "2025-11-06", "2025-11-10", False, "5 days"),
    ("    Coding", "2025-11-11", "2025-11-20", False, "10 days"),
    ("    Testing", "2025-11-21", "2025-11-25", False, "5 days"),
    
    ("Module 2: Job Descriptions", "2025-11-26", "2025-12-20", True, "1 mon"),
    ("    Planning", "2025-11-26", "2025-11-30", False, "5 days"),
    ("    Analysis and Design", "2025-12-01", "2025-12-05", False, "5 days"),
    ("    Coding", "2025-12-06", "2025-12-15", False, "10 days"),
    ("    Testing", "2025-12-16", "2025-12-20", False, "5 days"),
    
    ("Module 3: AI Resume Parsing", "2025-12-21", "2026-02-05", True, "1.5 mons"),
    ("    Planning", "2025-12-21", "2025-12-27", False, "7 days"),
    ("    Analysis and Design", "2025-12-28", "2026-01-03", False, "7 days"),
    ("    Coding", "2026-01-04", "2026-01-28", False, "25 days"),
    ("    Testing", "2026-01-29", "2026-02-05", False, "8 days"),

    ("Module 4: Interview Room", "2026-02-06", "2026-03-18", True, "1 mon"),
    ("    Planning", "2026-02-06", "2026-02-12", False, "7 days"),
    ("    Analysis and Design", "2026-02-13", "2026-02-19", False, "7 days"),
    ("    Coding", "2026-02-20", "2026-03-11", False, "20 days"),
    ("    Testing", "2026-03-12", "2026-03-18", False, "7 days"),
    
    ("Final Integration & Deployment", "2026-03-19", "2026-03-31", True, "13 days"),
    ("    System Testing", "2026-03-19", "2026-03-24", False, "6 days"),
    ("    Bug Fixing", "2026-03-25", "2026-03-28", False, "4 days"),
    ("    Deployment & Report", "2026-03-29", "2026-03-31", False, "3 days")
]

# Drawing the rows
fill_bar = PatternFill(start_color="71C1C4", end_color="71C1C4", fill_type="solid") # Cyan-like from screenshot
fill_header_bar = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")

row_idx = 3
for task in tasks:
    name, start_str, end_str, is_header, duration_str = task
    
    start_dt = datetime.strptime(start_str, "%Y-%m-%d")
    end_dt = datetime.strptime(end_str, "%Y-%m-%d")
    duration_days = (end_dt - start_dt).days + 1
    
    # Write details matching screenshot format
    ws.cell(row=row_idx, column=1, value=name)
    ws.cell(row=row_idx, column=2, value=duration_str)
    ws.cell(row=row_idx, column=3, value=start_dt.strftime("%a %d-%m-%y")) # E.g. Thu 28-11-25
    ws.cell(row=row_idx, column=4, value=end_dt.strftime("%a %d-%m-%y"))
    
    if is_header:
        for i in range(1, 5):
            ws.cell(row=row_idx, column=i).font = Font(bold=True)
            
    # Draw bars (color thin cells)
    start_offset = (start_dt - project_start).days
    for i in range(duration_days):
        col = 5 + start_offset + i
        if not is_header:
            ws.cell(row=row_idx, column=col).fill = fill_bar
        else:
            # Main module summary bar
            # In MS project it usually has a small black/grey line spanning the whole time
            # But making it cyan/grey works too. Let's make it light grey.
            ws.cell(row=row_idx, column=col).fill = fill_header_bar

    # For visual neatness, white out gridlines inside the bars by setting thin top/bottom borders of same color?
    # Simple solid fill is usually enough given how thin the columns are.

    row_idx += 1

# Save file
wb.save(r"e:\MSC-II\Technical_Recruiter Project\Technical_Recruiter_Gantt_Chart_v2.xlsx")
print("Saved v2 successfully!")
