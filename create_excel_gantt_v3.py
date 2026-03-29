import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt Chart (Aug 25 - Jan 26)"

# Set column widths for textual data
ws.column_dimensions['A'].width = 35  # Task Name
ws.column_dimensions['B'].width = 12  # Duration
ws.column_dimensions['C'].width = 18  # Start
ws.column_dimensions['D'].width = 18  # Finish

# Headers for textual data (Row 2, because Row 1 will have Month Headers)
headers = ["Task Name", "Duration", "Start", "Finish"]
for col, val in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col, value=val)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Define the project timeline (Aug 1 to Jan 31)
project_start = datetime(2025, 8, 1)
project_end = datetime(2026, 1, 31)
total_days = (project_end - project_start).days + 1

# Make the day columns VERY thin so it all fits on one screen (Width = 0.8)
for i in range(total_days):
    col = 5 + i
    col_letter = openpyxl.utils.get_column_letter(col)
    ws.column_dimensions[col_letter].width = 0.8

# Month grouping for Row 1
months = [
    ("Aug 2025", 31), ("Sep 2025", 30), ("Oct 2025", 31),
    ("Nov 2025", 30), ("Dec 2025", 31), ("Jan 2026", 31)
]

current_col = 5
for m_name, d_count in months:
    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + d_count - 1)
    cell = ws.cell(row=1, column=current_col, value=m_name)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    
    # Optional: Put faint lines separating months
    thin_border = Border(right=Side(style='thin', color="C0C0C0"))
    for r in range(1, 35):
        ws.cell(row=r, column=current_col + d_count - 1).border = thin_border
        
    current_col += d_count

# Tasks Data (Shifted to August Start)
tasks = [
    ("Technical Recruiter Platform", "2025-08-01", "2026-01-31", True, "6 mons"),
    ("Requirement Gathering", "2025-08-01", "2025-08-10", False, "10 days"),
    ("Planning", "2025-08-11", "2025-08-20", False, "10 days"),
    ("Analysis and Design", "2025-08-21", "2025-08-31", False, "10 days"),
    
    ("Module 1: Registration & Auth", "2025-09-01", "2025-09-25", True, "1 mon"),
    ("    Planning", "2025-09-01", "2025-09-05", False, "5 days"),
    ("    Analysis and Design", "2025-09-06", "2025-09-10", False, "5 days"),
    ("    Coding", "2025-09-11", "2025-09-20", False, "10 days"),
    ("    Testing", "2025-09-21", "2025-09-25", False, "5 days"),
    
    ("Module 2: Job Descriptions", "2025-09-26", "2025-10-20", True, "1 mon"),
    ("    Planning", "2025-09-26", "2025-09-30", False, "5 days"),
    ("    Analysis and Design", "2025-10-01", "2025-10-05", False, "5 days"),
    ("    Coding", "2025-10-06", "2025-10-15", False, "10 days"),
    ("    Testing", "2025-10-16", "2025-10-20", False, "5 days"),
    
    ("Module 3: AI Resume Parsing", "2025-10-21", "2025-12-05", True, "1.5 mons"),
    ("    Planning", "2025-10-21", "2025-10-27", False, "7 days"),
    ("    Analysis and Design", "2025-10-28", "2025-11-03", False, "7 days"),
    ("    Coding", "2025-11-04", "2025-11-28", False, "25 days"),
    ("    Testing", "2025-11-29", "2025-12-05", False, "7 days"),

    ("Module 4: Interview Room", "2025-12-06", "2026-01-18", True, "1 mon"),
    ("    Planning", "2025-12-06", "2025-12-12", False, "7 days"),
    ("    Analysis and Design", "2025-12-13", "2025-12-19", False, "7 days"),
    ("    Coding", "2025-12-20", "2026-01-11", False, "23 days"),
    ("    Testing", "2026-01-12", "2026-01-18", False, "7 days"),
    
    ("Final Integration & Deployment", "2026-01-19", "2026-01-31", True, "13 days"),
    ("    System Testing", "2026-01-19", "2026-01-24", False, "6 days"),
    ("    Bug Fixing", "2026-01-25", "2026-01-28", False, "4 days"),
    ("    Deployment & Report", "2026-01-29", "2026-01-31", False, "3 days")
]

# Drawing the rows
fill_bar = PatternFill(start_color="71C1C4", end_color="71C1C4", fill_type="solid") # Cyan-like from screenshot
fill_header_bar = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")

row_idx = 3
for task in tasks:
    name, start_str, end_str, is_header, duration_str = task
    
    start_dt = datetime.strptime(start_str, "%Y-%m-%d")
    end_dt = datetime.strptime(end_str, "%Y-%m-%d")
    duration_days = (end_dt - start_dt).days + 1
    
    # Write details matching screenshot format
    ws.cell(row=row_idx, column=1, value=name)
    ws.cell(row=row_idx, column=2, value=duration_str)
    ws.cell(row=row_idx, column=3, value=start_dt.strftime("%a %d-%m-%y")) # E.g. Thu 28-11-25
    ws.cell(row=row_idx, column=4, value=end_dt.strftime("%a %d-%m-%y"))
    
    if is_header:
        for i in range(1, 5):
            ws.cell(row=row_idx, column=i).font = Font(bold=True)
            
    # Draw bars (color thin cells)
    start_offset = (start_dt - project_start).days
    for i in range(duration_days):
        col = 5 + start_offset + i
        if not is_header:
            ws.cell(row=row_idx, column=col).fill = fill_bar
        else:
            ws.cell(row=row_idx, column=col).fill = fill_header_bar

    # For visual neatness, if not header, we can add top/bottom thin borders
    if not is_header:
        border_style = Border(top=Side(style='thin', color="FFFFFF"), bottom=Side(style='thin', color="FFFFFF"))
        for i in range(duration_days):
            ws.cell(row=row_idx, column=5 + start_offset + i).border = border_style

    row_idx += 1

# Save file
wb.save(r"e:\MSC-II\Technical_Recruiter Project\Technical_Recruiter_Gantt_Chart_v3.xlsx")
print("Saved v3 successfully!")
